from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.utils import timezone
from django.http import JsonResponse
from .models import Staff
from Blotter_Module.models import Blotter, Schedule, Evidence, BlotterAuditLog, Notification
from Blotter_Module.forms import ScheduleForm, EvidenceForm, CommentForm, BlotterStatusForm
from django.contrib.auth.models import User
from django.template.loader import get_template
from .models import Announcement
from .forms import AnnouncementForm
from .models import Appointment
from .forms import AppointmentForm, AppointmentStatusForm
from certificates.notification_utils import send_claim_notification, send_rejection_notification
from certificates.models import CertificateRequest
from django.http import JsonResponse
import json
# ====================== STAFF DASHBOARD ======================
@login_required(login_url='login')
def staff_dashboard(request):
    """Staff dashboard with statistics"""
    try:
        staff_profile = request.user.staff_profile
        
        # Get statistics from blotter module
        total_blotters = Blotter.objects.count()
        pending = Blotter.objects.filter(status='pending').count()
        scheduled = Schedule.objects.filter(is_completed=False).count()
        resolved = Blotter.objects.filter(status='resolved').count()
        unsettled = Blotter.objects.filter(status='in_progress').count()
        
        # NEW: Pending approvals count for staff review
        pending_approvals_count = Blotter.objects.filter(is_approved=False, status='pending').count()
        
        # Today's hearings
        today_hearings = Schedule.objects.filter(
            hearing_date__date=timezone.now().date(),
            is_completed=False
        ).select_related('blotter')[:5]
        
        # Recent blotters
        recent_blotters = Blotter.objects.all().order_by('-created_at')[:10]
        
        context = {
            'staff': staff_profile,
            'user': request.user,
            'total_blotters': total_blotters,
            'pending': pending,
            'scheduled': scheduled,
            'resolved': resolved,
            'unsettled': unsettled,
            'pending_approvals_count': pending_approvals_count,
            'today_hearings': today_hearings,
            'recent_blotters': recent_blotters,
        }
        return render(request, 'staff_module/dashboard.html', context)
    except Staff.DoesNotExist:
        messages.error(request, 'Access denied. Staff only area.')
        return redirect('dashboard')

@login_required(login_url='login')
def pending_approvals(request):
    """View all blotters pending staff approval"""
    try:
        staff_profile = request.user.staff_profile
        
        pending_blotters = Blotter.objects.filter(is_approved=False, status='pending').order_by('-created_at')
        
        # Search functionality
        search_query = request.GET.get('search', '')
        if search_query:
            pending_blotters = pending_blotters.filter(
                Q(blotter_number__icontains=search_query) |
                Q(complainant_name__icontains=search_query) |
                Q(respondent_name__icontains=search_query)
            )
        
        paginator = Paginator(pending_blotters, 20)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        context = {
            'staff': staff_profile,
            'blotters': page_obj,
            'search_query': search_query,
            'total_pending': pending_blotters.count(),
            'user': request.user,
        }
        return render(request, 'staff_module/pending_approvals.html', context)
    except Staff.DoesNotExist:
        messages.error(request, 'Access denied. Staff only area.')
        return redirect('dashboard')

# ====================== BLOTTER MANAGEMENT ======================
@login_required(login_url='login')
def staff_blotter_list(request):
    """View all blotters with advanced filters"""
    try:
        staff_profile = request.user.staff_profile
        blotters = Blotter.objects.all().order_by('-created_at')
        
        # Search
        search_query = request.GET.get('search', '')
        if search_query:
            blotters = blotters.filter(
                Q(blotter_number__icontains=search_query) |
                Q(complainant_name__icontains=search_query) |
                Q(respondent_name__icontains=search_query) |
                Q(complainant_phone__icontains=search_query)
            )
        
        # Status filter
        status_filter = request.GET.get('status', '')
        if status_filter:
            blotters = blotters.filter(status=status_filter)
        
        # Incident type filter
        incident_type_filter = request.GET.get('incident_type', '')
        if incident_type_filter:
            blotters = blotters.filter(incident_type=incident_type_filter)
        
        # Date range filter
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')
        if date_from:
            blotters = blotters.filter(created_at__date__gte=date_from)
        if date_to:
            blotters = blotters.filter(created_at__date__lte=date_to)
        
        # Verification method filter
        verification_filter = request.GET.get('verification_method', '')
        if verification_filter:
            blotters = blotters.filter(verification_method=verification_filter)
        
        # Sorting
        sort_by = request.GET.get('sort', '-created_at')
        allowed_sorts = ['created_at', '-created_at', 'blotter_number', '-blotter_number', 'status', '-status', 'incident_date', '-incident_date']
        if sort_by in allowed_sorts:
            blotters = blotters.order_by(sort_by)
        else:
            blotters = blotters.order_by('-created_at')
        
        # Pagination
        paginator = Paginator(blotters, 20)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        context = {
            'staff': staff_profile,
            'blotters': page_obj,
            'search_query': search_query,
            'status_filter': status_filter,
            'incident_type_filter': incident_type_filter,
            'date_from': date_from,
            'date_to': date_to,
            'verification_filter': verification_filter,
            'sort_by': sort_by,
            'status_choices': Blotter.STATUS_CHOICES,
            'incident_choices': Blotter.INCIDENT_CHOICES,
            'total_count': blotters.count(),
            'user': request.user,
        }
        return render(request, 'staff_module/blotter_list.html', context)
    except Staff.DoesNotExist:
        messages.error(request, 'Access denied. Staff only area.')
        return redirect('dashboard')


@login_required(login_url='login')
def staff_blotter_detail(request, blotter_id):
    """View blotter details"""
    try:
        staff_profile = request.user.staff_profile
        blotter = get_object_or_404(Blotter, id=blotter_id)
        schedules = blotter.schedules.all()
        evidences = blotter.evidences.all()  # This should work after migration
        
        # Get related appointment if exists
        appointment = Appointment.objects.filter(blotter=blotter, appointment_type='blotter_hearing').first()
        
        context = {
            'staff': staff_profile,
            'blotter': blotter,
            'schedules': schedules,
            'evidences': evidences,
            'appointment': appointment,
            'user': request.user,
        }
        return render(request, 'staff_module/blotter_detail.html', context)
    except Staff.DoesNotExist:
        messages.error(request, 'Access denied. Staff only area.')
        return redirect('dashboard')


@login_required(login_url='login')
def staff_schedule_hearing(request, blotter_id):
    """Schedule a hearing for a blotter and create appointment"""
    try:
        staff_profile = request.user.staff_profile
        blotter = get_object_or_404(Blotter, id=blotter_id)

        if request.method == 'POST':
            hearing_type = request.POST.get('hearing_type')
            hearing_date = request.POST.get('hearing_date')
            location = request.POST.get('location')
            notes = request.POST.get('notes', '')

            # Create schedule
            schedule = Schedule.objects.create(
                blotter=blotter,
                hearing_type=hearing_type,
                hearing_date=hearing_date,
                location=location,
                notes=notes,
                created_by=request.user,
                outcome='pending'
            )

            # Update blotter status
            blotter.status = 'scheduled'
            blotter.hearing_date = hearing_date
            blotter.save()

            # ===== CREATE APPOINTMENT FOR KAPITAN =====
            appointment = Appointment.objects.create(
                appointment_type='blotter_hearing',
                blotter=blotter,
                resident_name=blotter.complainant_name,
                resident_address=blotter.complainant_address,
                resident_contact=blotter.complainant_phone or '',
                resident_email=blotter.complainant_email or '',
                purpose=f"Blotter Hearing: {blotter.get_incident_type_display()} - {blotter.blotter_number}",
                appointment_date=hearing_date,
                duration_minutes=30,
                priority='important',
                hearing_type=hearing_type,
                location=location,
                status='approved',
                created_by=request.user,
                approved_by=request.user,
                approved_at=timezone.now(),
                notes=f"Schedule notes: {notes}"
            )

            messages.success(request, f'Hearing scheduled successfully! Appointment reference: {appointment.reference_number}')
            return redirect('staff_module:staff_blotter_detail', blotter_id=blotter.id)
        else:
            form = ScheduleForm()

        context = {
            'staff': staff_profile,
            'blotter': blotter,
            'form': form,
            'user': request.user,
        }
        return render(request, 'staff_module/schedule_hearing.html', context)
    except Staff.DoesNotExist:
        messages.error(request, 'Access denied. Staff only area.')
        return redirect('dashboard')


@login_required(login_url='login')
def staff_update_status(request, blotter_id):
    """Update blotter status"""
    try:
        staff_profile = request.user.staff_profile
        blotter = get_object_or_404(Blotter, id=blotter_id)
        old_status = blotter.status
        
        if request.method == 'POST':
            new_status = request.POST.get('status')
            reason = request.POST.get('reason', '')
            
            blotter.status = new_status
            if new_status == 'resolved':
                blotter.resolution_notes = reason
            blotter.save()
            
            messages.success(request, f'Status updated to {new_status}')
            return redirect('staff_module:staff_blotter_detail', blotter_id=blotter.id)
        
        context = {
            'staff': staff_profile,
            'blotter': blotter,
            'status_choices': Blotter.STATUS_CHOICES,
            'user': request.user,
        }
        return render(request, 'staff_module/update_status.html', context)
    except Staff.DoesNotExist:
        messages.error(request, 'Access denied. Staff only area.')
        return redirect('dashboard')


@login_required(login_url='login')
def staff_upload_evidence(request, blotter_id):
    """Upload evidence for a blotter"""
    try:
        staff_profile = request.user.staff_profile
        blotter = get_object_or_404(Blotter, id=blotter_id)
        
        if request.method == 'POST':
            form = EvidenceForm(request.POST, request.FILES)
            if form.is_valid():
                evidence = form.save(commit=False)
                evidence.blotter = blotter
                evidence.uploaded_by = request.user
                
                # Auto-detect file type
                if evidence.is_image():
                    evidence.file_type = 'image'
                else:
                    evidence.file_type = 'document'
                
                evidence.save()
                
                messages.success(request, 'Evidence uploaded successfully!')
                return redirect('staff_module:staff_blotter_detail', blotter_id=blotter.id)
        else:
            form = EvidenceForm()
        
        context = {
            'staff': staff_profile,
            'blotter': blotter,
            'form': form,
            'user': request.user,
        }
        return render(request, 'staff_module/upload_evidence.html', context)
    except Staff.DoesNotExist:
        messages.error(request, 'Access denied. Staff only area.')
        return redirect('dashboard')


# ====================== STAFF MANAGEMENT (ADMIN ONLY) ======================
@login_required(login_url='login')
def staff_list(request):
    """View all staff members (admin only)"""
    try:
        if request.user.staff_profile.role != 'admin':
            messages.error(request, 'Admin access required')
            return redirect('staff_module:staff_dashboard')
        
        staff_members = Staff.objects.all().select_related('user')
        
        context = {
            'staff_members': staff_members,
            'staff': request.user.staff_profile,
            'user': request.user,
        }
        return render(request, 'staff_module/staff_list.html', context)
    except Staff.DoesNotExist:
        messages.error(request, 'Access denied. Staff only area.')
        return redirect('dashboard')


@login_required(login_url='login')
def staff_create(request):
    """Create new staff member (admin only)"""
    try:
        if request.user.staff_profile.role != 'admin':
            messages.error(request, 'Admin access required')
            return redirect('staff_module:staff_dashboard')
        
        if request.method == 'POST':
            username = request.POST.get('username')
            email = request.POST.get('email')
            password = request.POST.get('password')
            role = request.POST.get('role')
            position = request.POST.get('position')
            contact_number = request.POST.get('contact_number')
            first_name = request.POST.get('first_name', '')
            last_name = request.POST.get('last_name', '')
            
            # Check if username exists
            if User.objects.filter(username=username).exists():
                messages.error(request, 'Username already exists')
                return redirect('staff_module:staff_create')
            
            # Create user
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name
            )
            
            # Update staff profile (created by signal)
            user.staff_profile.role = role
            user.staff_profile.position = position
            user.staff_profile.contact_number = contact_number
            user.staff_profile.save()
            
            messages.success(request, f'Staff {username} created successfully!')
            return redirect('staff_module:staff_list')
        
        context = {
            'staff': request.user.staff_profile,
            'user': request.user,
            'role_choices': Staff.ROLE_CHOICES,
        }
        return render(request, 'staff_module/staff_create.html', context)
    except Staff.DoesNotExist:
        messages.error(request, 'Access denied. Staff only area.')
        return redirect('dashboard')


@login_required(login_url='login')
def staff_edit(request, staff_id):
    """Edit staff member (admin only)"""
    try:
        if request.user.staff_profile.role != 'admin':
            messages.error(request, 'Admin access required')
            return redirect('staff_module:staff_dashboard')
        
        staff = get_object_or_404(Staff, id=staff_id)
        
        if request.method == 'POST':
            staff.role = request.POST.get('role')
            staff.position = request.POST.get('position')
            staff.contact_number = request.POST.get('contact_number')
            staff.address = request.POST.get('address', '')
            staff.is_active = request.POST.get('is_active') == 'on'
            staff.save()
            
            # Update user info
            staff.user.first_name = request.POST.get('first_name', '')
            staff.user.last_name = request.POST.get('last_name', '')
            staff.user.email = request.POST.get('email', '')
            staff.user.save()
            
            messages.success(request, f'Staff {staff.user.username} updated successfully!')
            return redirect('staff_module:staff_list')
        
        context = {
            'staff': staff,
            'user_staff': request.user.staff_profile,
            'user': request.user,
            'role_choices': Staff.ROLE_CHOICES,
        }
        return render(request, 'staff_module/staff_edit.html', context)
    except Staff.DoesNotExist:
        messages.error(request, 'Access denied. Staff only area.')
        return redirect('dashboard')


@login_required(login_url='login')
def staff_delete(request, staff_id):
    """Delete staff member (admin only)"""
    try:
        if request.user.staff_profile.role != 'admin':
            messages.error(request, 'Admin access required')
            return redirect('staff_module:staff_dashboard')
        
        staff = get_object_or_404(Staff, id=staff_id)
        
        # Prevent deleting yourself
        if staff.user.id == request.user.id:
            messages.error(request, 'You cannot delete your own account!')
            return redirect('staff_module:staff_list')
        
        if request.method == 'POST':
            username = staff.user.username
            staff.user.delete()  # Staff profile will be deleted due to CASCADE
            messages.success(request, f'Staff {username} has been deleted successfully!')
            return redirect('staff_module:staff_list')
        
        context = {
            'staff': staff,
            'user_staff': request.user.staff_profile,
            'user': request.user,
        }
        return render(request, 'staff_module/staff_confirm_delete.html', context)
    except Staff.DoesNotExist:
        messages.error(request, 'Access denied. Staff only area.')
        return redirect('dashboard')


@login_required(login_url='login')
def staff_generate_summon(request, blotter_id):
    """Generate summon letter for respondent"""
    try:
        staff_profile = request.user.staff_profile
        blotter = get_object_or_404(Blotter, id=blotter_id)
        
        # Check if hearing is scheduled
        latest_schedule = blotter.schedules.filter(is_completed=False).order_by('-hearing_date').first()
        
        if not latest_schedule:
            messages.error(request, f'Cannot generate summon letter. Please schedule a hearing first for Blotter #{blotter.blotter_number}.')
            return redirect('staff_module:staff_blotter_detail', blotter_id=blotter.id)
        
        # Use the schedule date instead of blotter.hearing_date
        hearing_date = latest_schedule.hearing_date
        
        context = {
            'blotter': blotter,
            'staff': staff_profile,
            'today': timezone.now(),
            'hearing_date': hearing_date,
            'schedule': latest_schedule,
        }
        
        return render(request, 'staff_module/summon_letter.html', context)
        
    except Staff.DoesNotExist:
        messages.error(request, 'Access denied. Staff only area.')
        return redirect('dashboard')


@login_required(login_url='login')
def staff_update_hearing_status(request, hearing_id):
    """Update hearing/mediation session status"""
    try:
        staff_profile = request.user.staff_profile
        hearing = get_object_or_404(Schedule, id=hearing_id)
        blotter = hearing.blotter
        
        if request.method == 'POST':
            attendance = request.POST.get('attendance')
            outcome = request.POST.get('outcome')
            outcome_notes = request.POST.get('outcome_notes', '')
            
            # Update hearing record
            hearing.attendance = attendance
            hearing.outcome = outcome
            hearing.outcome_notes = outcome_notes
            hearing.is_completed = outcome in ['completed', 'settled', 'unsettled']
            hearing.save()
            
            # Update related appointment status
            appointment = Appointment.objects.filter(blotter=blotter, appointment_type='blotter_hearing').first()
            if appointment:
                if outcome == 'settled':
                    appointment.status = 'completed'
                elif outcome == 'unsettled':
                    appointment.status = 'cancelled'
                else:
                    appointment.status = 'completed'
                appointment.notes = f"Hearing outcome: {outcome}. Notes: {outcome_notes}"
                appointment.save()
            
            # Update blotter status based on hearing outcome
            if outcome == 'settled':
                blotter.status = 'resolved'
                blotter.resolution_notes = outcome_notes
            elif outcome == 'unsettled':
                blotter.status = 'unsettled'
            elif outcome == 'failed_complainant' or outcome == 'failed_respondent':
                blotter.status = 'dismissed'
            else:
                blotter.status = 'in_progress'
            
            blotter.save()
            
            # Create audit log
            BlotterAuditLog.objects.create(
                blotter=blotter,
                action='status_change',
                old_value=hearing.get_outcome_display(),
                new_value=outcome,
                performed_by=request.user
            )
            
            messages.success(request, f'Hearing status updated successfully!')
            return redirect('staff_module:staff_blotter_detail', blotter_id=blotter.id)
        
        context = {
            'staff': staff_profile,
            'hearing': hearing,
            'blotter': blotter,
            'user': request.user,
            'attendance_choices': Schedule.ATTENDANCE_CHOICES,
            'outcome_choices': Schedule.OUTCOME_CHOICES,
        }
        return render(request, 'staff_module/update_hearing_status.html', context)
        
    except Staff.DoesNotExist:
        messages.error(request, 'Access denied. Staff only area.')
        return redirect('dashboard')


@login_required(login_url='login')
def staff_pending_approvals(request):
    """View all blotters pending staff approval"""
    try:
        staff_profile = request.user.staff_profile

        pending_blotters = Blotter.objects.filter(is_approved=False).order_by('-created_at')

        search_query = request.GET.get('search', '').strip()
        if search_query:
            pending_blotters = pending_blotters.filter(
                Q(blotter_number__icontains=search_query) |
                Q(complainant_name__icontains=search_query) |
                Q(respondent_name__icontains=search_query)
            )

        total_pending = pending_blotters.count()

        paginator = Paginator(pending_blotters, 20)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context = {
            'staff': staff_profile,
            'blotters': page_obj,
            'search_query': search_query,
            'total_pending': total_pending,
            'user': request.user,
        }
        return render(request, 'staff_module/pending_approvals.html', context)
    except Staff.DoesNotExist:
        messages.error(request, 'Access denied. Staff only area.')
        return redirect('dashboard')


@login_required(login_url='login')
def staff_approve_blotter(request, blotter_id):
    """Approve a blotter complaint"""
    try:
        staff_profile = request.user.staff_profile
        blotter = get_object_or_404(Blotter, id=blotter_id)
        
        if request.method == 'POST':
            action = request.POST.get('action')
            
            if action == 'approve':
                blotter.is_approved = True
                blotter.approved_by = request.user
                blotter.approved_at = timezone.now()
                blotter.verification_status = 'approved'
                messages.success(request, f'Blotter {blotter.blotter_number} has been approved.')
                
            elif action == 'reject':
                rejection_reason = request.POST.get('rejection_reason', '')
                blotter.is_approved = False
                blotter.rejection_reason = rejection_reason
                blotter.status = 'dismissed'
                messages.warning(request, f'Blotter {blotter.blotter_number} has been rejected.')
            
            blotter.reviewed_by = request.user
            blotter.reviewed_at = timezone.now()
            blotter.save()
            
            return redirect('staff_module:staff_pending_approvals')
        
        context = {
            'staff': staff_profile,
            'blotter': blotter,
            'user': request.user,
        }
        return render(request, 'staff_module/approve_blotter.html', context)
    except Staff.DoesNotExist:
        messages.error(request, 'Access denied. Staff only area.')
        return redirect('dashboard')


# ====================== ANNOUNCEMENTS ======================
@login_required
def announcement_list(request):
    """View all announcements (Staff)"""
    announcements = Announcement.objects.all().order_by('-created_at')
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter == 'active':
        announcements = announcements.filter(is_active=True, expires_at__isnull=True) | announcements.filter(expires_at__gt=timezone.now())
    elif status_filter == 'expired':
        announcements = announcements.filter(expires_at__lt=timezone.now())
    
    paginator = Paginator(announcements, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'announcements': page_obj,
        'status_filter': status_filter,
    }
    return render(request, 'staff_module/announcement_list.html', context)


@login_required
def announcement_create(request):
    """Create new announcement"""
    if request.method == 'POST':
        form = AnnouncementForm(request.POST, request.FILES)
        if form.is_valid():
            announcement = form.save(commit=False)
            announcement.created_by = request.user
            announcement.save()
            messages.success(request, 'Announcement posted successfully!')
            return redirect('staff_module:announcement_list')
    else:
        form = AnnouncementForm()
    
    context = {
        'form': form,
        'is_edit': False,
    }
    return render(request, 'staff_module/announcement_form.html', context)


@login_required
def announcement_edit(request, pk):
    """Edit announcement"""
    announcement = get_object_or_404(Announcement, pk=pk)
    
    if request.method == 'POST':
        form = AnnouncementForm(request.POST, request.FILES, instance=announcement)
        if form.is_valid():
            form.save()
            messages.success(request, 'Announcement updated successfully!')
            return redirect('staff_module:announcement_list')
    else:
        form = AnnouncementForm(instance=announcement)
    
    context = {
        'form': form,
        'announcement': announcement,
        'is_edit': True,
    }
    return render(request, 'staff_module/announcement_form.html', context)


@login_required
def announcement_delete(request, pk):
    """Delete announcement"""
    announcement = get_object_or_404(Announcement, pk=pk)
    if request.method == 'POST':
        announcement.delete()
        messages.success(request, 'Announcement deleted successfully!')
        return redirect('staff_module:announcement_list')
    
    return render(request, 'staff_module/announcement_confirm_delete.html', {'announcement': announcement})


@login_required
def announcement_toggle_status(request, pk):
    """Toggle announcement active status"""
    announcement = get_object_or_404(Announcement, pk=pk)
    announcement.is_active = not announcement.is_active
    announcement.save()
    status = 'activated' if announcement.is_active else 'deactivated'
    messages.success(request, f'Announcement {status} successfully!')
    return redirect('staff_module:announcement_list')


# ====================== APPOINTMENTS ======================
@login_required
def appointment_list(request):
    """View all appointments (includes blotter hearings)"""
    appointments = Appointment.objects.all().order_by('-created_at')
    
    # Filter by type
    type_filter = request.GET.get('type', '')
    if type_filter:
        appointments = appointments.filter(appointment_type=type_filter)
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        appointments = appointments.filter(status=status_filter)
    
    # Filter by date
    date_filter = request.GET.get('date', '')
    if date_filter == 'today':
        appointments = appointments.filter(appointment_date__date=timezone.now().date())
    elif date_filter == 'upcoming':
        appointments = appointments.filter(appointment_date__gte=timezone.now(), status__in=['pending', 'approved'])
    
    paginator = Paginator(appointments, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'appointments': page_obj,
        'type_filter': type_filter,
        'status_filter': status_filter,
        'date_filter': date_filter,
        'status_choices': Appointment.STATUS_CHOICES,
    }
    return render(request, 'staff_module/appointment_list.html', context)


@login_required
def appointment_create(request):
    """Create new appointment"""
    if request.method == 'POST':
        form = AppointmentForm(request.POST)
        if form.is_valid():
            appointment = form.save(commit=False)
            appointment.created_by = request.user
            appointment.save()
            messages.success(request, f'Appointment created successfully! Reference: {appointment.reference_number}')
            return redirect('staff_module:appointment_list')
    else:
        form = AppointmentForm()
    
    context = {'form': form}
    return render(request, 'staff_module/appointment_form.html', context)


@login_required
def appointment_detail(request, pk):
    """View appointment details"""
    appointment = get_object_or_404(Appointment, pk=pk)
    context = {'appointment': appointment}
    return render(request, 'staff_module/appointment_detail.html', context)


@login_required
def appointment_update_status(request, pk):
    """Update appointment status"""
    appointment = get_object_or_404(Appointment, pk=pk)
    
    if request.method == 'POST':
        form = AppointmentStatusForm(request.POST)
        if form.is_valid():
            new_status = form.cleaned_data['status']
            notes = form.cleaned_data['notes']
            
            appointment.status = new_status
            if notes:
                appointment.notes = notes
            if new_status == 'approved':
                appointment.approved_by = request.user
                appointment.approved_at = timezone.now()
            appointment.save()
            
            messages.success(request, f'Appointment {appointment.reference_number} status updated to {new_status}')
            return redirect('staff_module:appointment_detail', pk=appointment.id)
    else:
        form = AppointmentStatusForm(initial={'status': appointment.status})
    
    context = {
        'appointment': appointment,
        'form': form,
    }
    return render(request, 'staff_module/appointment_status.html', context)


@login_required
def appointment_calendar(request):
    """Calendar view of appointments"""
    appointments = Appointment.objects.filter(
        status__in=['pending', 'approved'],
        appointment_date__gte=timezone.now()
    ).order_by('appointment_date')
    
    # Prepare data for JavaScript
    appointments_json = []
    for apt in appointments:
        appointments_json.append({
            'id': apt.id,
            'date': apt.appointment_date.strftime('%Y-%m-%d'),
            'time': apt.appointment_date.strftime('%I:%M %p'),
            'resident_name': apt.resident_name,
            'purpose': apt.purpose,
            'reference_number': apt.reference_number,
        })
    
    import json
    context = {
        'appointments_json': json.dumps(appointments_json),
        'current_year': timezone.now().year,
        'current_month': timezone.now().month,
        'current_month_name': timezone.now().strftime('%B'),
    }
    return render(request, 'staff_module/appointment_calendar.html', context)


@login_required
def release_request(request, request_id):
    """Staff view: Release certificate after Kapitan approval"""
    cert_request = get_object_or_404(CertificateRequest, id=request_id)
    
    if request.method == 'POST':
        if cert_request.status != 'approved':
            messages.error(request, f'Cannot release certificate with status: {cert_request.get_status_display()}. Only approved certificates can be released.')
            return redirect('certificates:view_request', request_id=cert_request.id)
        
        cert_request.status = 'released'
        cert_request.date_released = timezone.now()
        cert_request.remarks = request.POST.get('remarks', '')
        cert_request.save()
        
        # Send claim notification if not already sent
        send_claim_notification(cert_request)
        
        messages.success(request, f'Certificate {cert_request.request_id} has been marked as released.')
        return redirect('certificates:request_list')
    
    context = {
        'request': cert_request,
        'type_display': dict(CertificateRequest.CERTIFICATE_TYPES).get(cert_request.request_type, ''),
    }
    return render(request, 'certificates/release_request.html', context)


# ====================== REPORTS VIEWS ======================

@login_required
def reports_dashboard(request):
    """Staff reports dashboard with various report options"""
    try:
        staff_profile = request.user.staff_profile
        
        # Get date range for filtering
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')
        
        # Default to current month if no dates provided
        if not date_from:
            date_from = timezone.now().replace(day=1).strftime('%Y-%m-%d')
        if not date_to:
            date_to = timezone.now().strftime('%Y-%m-%d')
        
        context = {
            'staff': staff_profile,
            'user': request.user,
            'date_from': date_from,
            'date_to': date_to,
        }
        return render(request, 'staff_module/reports/reports_dashboard.html', context)
        
    except Staff.DoesNotExist:
        messages.error(request, 'Access denied. Staff only area.')
        return redirect('dashboard')


@login_required
def blotter_report(request):
    """Generate blotter report"""
    try:
        staff_profile = request.user.staff_profile
        
        # Get filters
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')
        status_filter = request.GET.get('status', '')
        incident_type = request.GET.get('incident_type', '')
        
        # Base queryset
        blotters = Blotter.objects.filter(is_approved=True)
        
        # Apply filters
        if date_from:
            blotters = blotters.filter(created_at__date__gte=date_from)
        if date_to:
            blotters = blotters.filter(created_at__date__lte=date_to)
        if status_filter:
            blotters = blotters.filter(status=status_filter)
        if incident_type:
            blotters = blotters.filter(incident_type=incident_type)
        
        # Statistics
        total_blotters = blotters.count()
        pending_count = blotters.filter(status='pending').count()
        scheduled_count = blotters.filter(status='scheduled').count()
        resolved_count = blotters.filter(status='resolved').count()
        in_progress_count = blotters.filter(status='in_progress').count()
        unsettled_count = blotters.filter(status='unsettled').count()
        
        # Group by incident type
        incident_stats = []
        for incident_code, incident_label in Blotter.INCIDENT_CHOICES:
            count = blotters.filter(incident_type=incident_code).count()
            if count > 0:
                incident_stats.append({
                    'type': incident_label,
                    'count': count,
                    'percentage': round((count / total_blotters * 100), 1) if total_blotters > 0 else 0
                })
        
        # Group by month
        monthly_stats = blotters.extra(
            select={'month': "DATE_FORMAT(created_at, '%%Y-%%m')"}
        ).values('month').annotate(count=Count('id')).order_by('month')
        
        # Export to Excel if requested
        if request.GET.get('export') == 'excel':
            return export_blotter_excel(blotters, date_from, date_to)
        
        context = {
            'staff': staff_profile,
            'user': request.user,
            'report_type': 'blotter',
            'blotters': blotters[:100],  # Limit to 100 for display
            'total_blotters': total_blotters,
            'pending_count': pending_count,
            'scheduled_count': scheduled_count,
            'resolved_count': resolved_count,
            'in_progress_count': in_progress_count,
            'unsettled_count': unsettled_count,
            'incident_stats': incident_stats,
            'monthly_stats': monthly_stats,
            'date_from': date_from,
            'date_to': date_to,
            'status_filter': status_filter,
            'incident_type': incident_type,
            'status_choices': Blotter.STATUS_CHOICES,
            'incident_choices': Blotter.INCIDENT_CHOICES,
        }
        return render(request, 'staff_module/reports/report_blotter.html', context)
        
    except Staff.DoesNotExist:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')


@login_required
def certificate_report(request):
    """Generate certificate requests report"""
    try:
        staff_profile = request.user.staff_profile
        
        # Get filters
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')
        status_filter = request.GET.get('status', '')
        cert_type = request.GET.get('cert_type', '')
        
        # Base queryset
        certificates = CertificateRequest.objects.all()
        
        # Apply filters
        if date_from:
            certificates = certificates.filter(date_submitted__date__gte=date_from)
        if date_to:
            certificates = certificates.filter(date_submitted__date__lte=date_to)
        if status_filter:
            certificates = certificates.filter(status=status_filter)
        if cert_type:
            certificates = certificates.filter(request_type=cert_type)
        
        # Statistics
        total_requests = certificates.count()
        pending_count = certificates.filter(status='pending').count()
        for_approval_count = certificates.filter(status='for_approval').count()
        approved_count = certificates.filter(status='approved').count()
        rejected_count = certificates.filter(status='rejected').count()
        released_count = certificates.filter(status='released').count()
        
        # Group by certificate type
        cert_type_stats = []
        type_map = {
            'clearance': 'Barangay Clearance',
            'indigency': 'Certificate of Indigency',
            'residency': 'Certificate of Residency',
            'id': 'Barangay ID',
        }
        for type_code, type_label in type_map.items():
            count = certificates.filter(request_type=type_code).count()
            if count > 0:
                cert_type_stats.append({
                    'type': type_label,
                    'count': count,
                    'percentage': round((count / total_requests * 100), 1) if total_requests > 0 else 0
                })
        
        # Monthly trend
        monthly_stats = certificates.extra(
            select={'month': "DATE_FORMAT(date_submitted, '%%Y-%%m')"}
        ).values('month').annotate(count=Count('id')).order_by('month')
        
        # Export to Excel if requested
        if request.GET.get('export') == 'excel':
            return export_certificate_excel(certificates, date_from, date_to)
        
        context = {
            'staff': staff_profile,
            'user': request.user,
            'report_type': 'certificate',
            'certificates': certificates[:100],
            'total_requests': total_requests,
            'pending_count': pending_count,
            'for_approval_count': for_approval_count,
            'approved_count': approved_count,
            'rejected_count': rejected_count,
            'released_count': released_count,
            'cert_type_stats': cert_type_stats,
            'monthly_stats': monthly_stats,
            'date_from': date_from,
            'date_to': date_to,
            'status_filter': status_filter,
            'cert_type': cert_type,
            'status_choices': CertificateRequest.STATUS_CHOICES,
            'certificate_types': [
                ('clearance', 'Barangay Clearance'),
                ('indigency', 'Certificate of Indigency'),
                ('residency', 'Certificate of Residency'),
                ('id', 'Barangay ID'),
            ],
        }
        return render(request, 'staff_module/reports/report_certificate.html', context)
        
    except Staff.DoesNotExist:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')


@login_required
def hearing_report(request):
    """Generate hearing/mediation report"""
    try:
        staff_profile = request.user.staff_profile
        
        # Get filters
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')
        outcome_filter = request.GET.get('outcome', '')
        
        # Base queryset
        hearings = Schedule.objects.all()
        
        # Apply filters
        if date_from:
            hearings = hearings.filter(hearing_date__date__gte=date_from)
        if date_to:
            hearings = hearings.filter(hearing_date__date__lte=date_to)
        if outcome_filter:
            hearings = hearings.filter(outcome=outcome_filter)
        
        # Statistics
        total_hearings = hearings.count()
        completed_count = hearings.filter(is_completed=True).count()
        pending_count = hearings.filter(is_completed=False).count()
        
        # Settled count
        settled_hearings = hearings.filter(outcome='settled').count()
        unsettled_hearings = hearings.filter(outcome='unsettled').count()
        
        # Outcome statistics
        outcome_stats = []
        for outcome_code, outcome_label in Schedule.OUTCOME_CHOICES:
            count = hearings.filter(outcome=outcome_code).count()
            if count > 0:
                outcome_stats.append({
                    'outcome': outcome_label,
                    'count': count,
                    'percentage': round((count / total_hearings * 100), 1) if total_hearings > 0 else 0
                })
        
        context = {
            'staff': staff_profile,
            'user': request.user,
            'report_type': 'hearing',
            'hearings': hearings[:100],
            'total_hearings': total_hearings,
            'completed_count': completed_count,
            'pending_count': pending_count,
            'settled_hearings': settled_hearings,
            'unsettled_hearings': unsettled_hearings,
            'outcome_stats': outcome_stats,
            'date_from': date_from,
            'date_to': date_to,
            'outcome_filter': outcome_filter,
            'outcome_choices': Schedule.OUTCOME_CHOICES,
        }
        return render(request, 'staff_module/reports/report_hearing.html', context)
        
    except Staff.DoesNotExist:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')


@login_required
def summary_report(request):
    """Generate summary dashboard report"""
    try:
        staff_profile = request.user.staff_profile
        
        # Get date range
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')
        
        # Default to current month
        if not date_from:
            date_from = timezone.now().replace(day=1).strftime('%Y-%m-%d')
        if not date_to:
            date_to = timezone.now().strftime('%Y-%m-%d')
        
        # Blotter statistics
        blotters = Blotter.objects.filter(created_at__date__gte=date_from, created_at__date__lte=date_to)
        total_blotters = blotters.count()
        resolved_blotters = blotters.filter(status='resolved').count()
        pending_blotters = blotters.filter(status='pending').count()
        
        # Certificate statistics
        certificates = CertificateRequest.objects.filter(date_submitted__date__gte=date_from, date_submitted__date__lte=date_to)
        total_certificates = certificates.count()
        approved_certificates = certificates.filter(status='approved').count()
        for_approval_certificates = certificates.filter(status='for_approval').count()
        
        # Hearing statistics
        hearings = Schedule.objects.filter(hearing_date__date__gte=date_from, hearing_date__date__lte=date_to)
        total_hearings = hearings.count()
        settled_hearings = hearings.filter(outcome='settled').count()
        unsettled_hearings = hearings.filter(outcome='unsettled').count()
        
        context = {
            'staff': staff_profile,
            'user': request.user,
            'date_from': date_from,
            'date_to': date_to,
            'total_blotters': total_blotters,
            'resolved_blotters': resolved_blotters,
            'pending_blotters': pending_blotters,
            'total_certificates': total_certificates,
            'approved_certificates': approved_certificates,
            'for_approval_certificates': for_approval_certificates,
            'total_hearings': total_hearings,
            'settled_hearings': settled_hearings,
            'unsettled_hearings': unsettled_hearings,
        }
        return render(request, 'staff_module/reports/report_summary.html', context)
        
    except Staff.DoesNotExist:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')


# ====================== EXPORT FUNCTIONS ======================

def export_blotter_excel(blotters, date_from, date_to):
    """Export blotter data to Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Blotter Report"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['Blotter Number', 'Complainant', 'Respondent', 'Incident Type', 'Incident Date', 'Status', 'Date Filed', 'Resolution']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Data
    for row, blotter in enumerate(blotters, 2):
        ws.cell(row=row, column=1, value=blotter.blotter_number)
        ws.cell(row=row, column=2, value=blotter.complainant_name)
        ws.cell(row=row, column=3, value=blotter.respondent_name or 'N/A')
        ws.cell(row=row, column=4, value=blotter.get_incident_type_display())
        ws.cell(row=row, column=5, value=blotter.incident_date.strftime('%Y-%m-%d') if blotter.incident_date else '')
        ws.cell(row=row, column=6, value=blotter.get_status_display())
        ws.cell(row=row, column=7, value=blotter.created_at.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=8, value=blotter.resolution_notes or '')
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="blotter_report_{date_from}_to_{date_to}.xlsx"'
    wb.save(response)
    return response


def export_certificate_excel(certificates, date_from, date_to):
    """Export certificate data to Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Certificate Report"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    
    type_map = {
        'clearance': 'Barangay Clearance',
        'indigency': 'Certificate of Indigency',
        'residency': 'Certificate of Residency',
        'id': 'Barangay ID',
    }
    
    # Headers
    headers = ['Request ID', 'Resident Name', 'Certificate Type', 'Date Submitted', 'Status', 'Date Approved', 'Remarks']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row, cert in enumerate(certificates, 2):
        ws.cell(row=row, column=1, value=cert.request_id)
        ws.cell(row=row, column=2, value=cert.full_name)
        ws.cell(row=row, column=3, value=type_map.get(cert.request_type, cert.request_type))
        ws.cell(row=row, column=4, value=cert.date_submitted.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=5, value=cert.get_status_display())
        ws.cell(row=row, column=6, value=cert.date_approved.strftime('%Y-%m-%d') if cert.date_approved else '')
        ws.cell(row=row, column=7, value=cert.remarks or '')
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="certificate_report_{date_from}_to_{date_to}.xlsx"'
    wb.save(response)
    return response

@login_required(login_url='login')
def staff_approve_blotter_action(request, blotter_id):
    """Approve or reject a blotter via AJAX"""
    try:
        staff_profile = request.user.staff_profile
        blotter = get_object_or_404(Blotter, id=blotter_id)
        
        if request.method == 'POST':
            data = json.loads(request.body)
            action = data.get('action')
            
            if action == 'approve':
                blotter.is_approved = True
                blotter.approved_by = request.user
                blotter.approved_at = timezone.now()
                blotter.verification_status = 'approved'
                blotter.status = 'pending'
                blotter.save()
                
                # Create audit log
                BlotterAuditLog.objects.create(
                    blotter=blotter,
                    action='approve',
                    new_value=f'Blotter approved by {request.user.username}',
                    performed_by=request.user
                )
                
                messages.success(request, f'Blotter {blotter.blotter_number} has been approved.')
                
                return JsonResponse({
                    'success': True,
                    'message': f'Blotter {blotter.blotter_number} approved successfully!',
                    'redirect_url': request.META.get('HTTP_REFERER', '/staff/pending-approvals/')
                })
                
            elif action == 'reject':
                rejection_reason = data.get('rejection_reason', '')
                blotter.is_approved = False
                blotter.rejection_reason = rejection_reason
                blotter.status = 'dismissed'
                blotter.verification_status = 'rejected'
                blotter.save()
                
                # Create audit log
                BlotterAuditLog.objects.create(
                    blotter=blotter,
                    action='reject',
                    new_value=f'Blotter rejected by {request.user.username}. Reason: {rejection_reason}',
                    performed_by=request.user
                )
                
                messages.warning(request, f'Blotter {blotter.blotter_number} has been rejected.')
                
                return JsonResponse({
                    'success': True,
                    'message': f'Blotter {blotter.blotter_number} rejected.',
                    'redirect_url': '/staff/pending-approvals/'
                })
            
        return JsonResponse({'error': 'Invalid request method'}, status=400)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)